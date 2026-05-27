"""
CheXbert-based Clinical Accuracy Evaluation
=============================================
Evaluate LLM-generated radiology reports using CheXbert to extract
14-class CheXpert labels, then compare against ground truth labels.

Workflow:
  1. Generate two CSV files with "Report Impression" column:
     - Ground truth reports (from annotation.json)
     - Predicted reports (from LLM output JSON)
  2. Call CheXbert label.py on both CSVs to get labeled_reports.csv
  3. Read both labeled results and compute metrics (AUC, F1, Recall, Specificity)

CheXbert label.py output convention:
  - NaN  -> Blank (not mentioned)
  - 1    -> Positive
  - 0    -> Negative
  - -1   -> Uncertain

Binary strategy (U-zeros):
  - Only val == 1.0 is positive (1.0)
  - Everything else (NaN, 0, -1) is negative (0.0)

Usage:
    python evaluate_chexbert.py \
        --llm_output ./results/qwen_output.json \
        --annotation_json ./data/annotation.json \
        --chexbert_checkpoint ./checkpoints/chexbert.pth \
        --output_dir ./results/chexbert_eval
"""

import os
import sys
import json
import shutil
import tempfile
import argparse
import subprocess
import warnings

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

warnings.filterwarnings("ignore")

# ============================================================
# Constants
# ============================================================
CHEXPERT_LABELS = [
    "Enlarged Cardiomediastinum",
    "Cardiomegaly",
    "Lung Opacity",
    "Lung Lesion",
    "Edema",
    "Consolidation",
    "Pneumonia",
    "Atelectasis",
    "Pneumothorax",
    "Pleural Effusion",
    "Pleural Other",
    "Fracture",
    "Support Devices",
    "No Finding",
]


# ============================================================
# Step 1: Load LLM output
# ============================================================
def load_llm_output(json_path):
    """
    Load LLM-generated reports from JSON file.

    Expected format:
        {"test": [{"id": "dicom_id", "output": "report text"}, ...]}

    Returns:
        list of dicts with keys 'id' and 'output'
    """
    print(f"[INFO] Loading LLM output from: {json_path}")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    samples = data.get("test", data if isinstance(data, list) else [])
    print(f"[INFO] Loaded {len(samples)} LLM-generated reports")
    return samples


# ============================================================
# Step 2: Generate CSV files
# ============================================================
def create_prediction_csv(llm_samples, output_path):
    """Create a CSV file with 'Report Impression' column for LLM predictions."""
    reports = []
    for sample in llm_samples:
        text = sample.get("output", "")
        text = " ".join(text.split())
        reports.append(text)

    df = pd.DataFrame({"Report Impression": reports})
    df.to_csv(output_path, index=False)
    print(f"[INFO] Created prediction CSV: {output_path} ({len(reports)} reports)")
    return output_path


def create_ground_truth_csv(annotation_json_path, dicom_ids, output_path):
    """Create a CSV file with 'Report Impression' column for ground truth reports."""
    print(f"[INFO] Loading annotation.json for ground truth reports...")
    with open(annotation_json_path, "r", encoding="utf-8") as f:
        annotation = json.load(f)

    dicom_to_report = {}
    for split_name in ["train", "val", "test"]:
        if split_name not in annotation:
            continue
        for sample in annotation[split_name]:
            did = sample["id"]
            dicom_to_report[did] = sample.get("report", "")

    print(f"[INFO] Built dicom_id -> report mapping for {len(dicom_to_report)} samples")

    reports = []
    n_missing = 0
    for did in dicom_ids:
        report = dicom_to_report.get(did, "")
        if not report:
            n_missing += 1
        reports.append(report)

    df = pd.DataFrame({"Report Impression": reports})
    df.to_csv(output_path, index=False)
    print(f"[INFO] Created ground truth CSV: {output_path} ({len(reports)} reports)")
    if n_missing > 0:
        print(f"[WARN] {n_missing}/{len(reports)} dicom_ids had no report in annotation.json")
    return output_path


# ============================================================
# Step 3: Call CheXbert label.py
# ============================================================
def run_chexbert_labeling(csv_path, output_dir, checkpoint_path):
    """Call CheXbert label.py via subprocess to label reports."""
    label_script = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "CheXbert", "src", "label.py"
    )

    if not os.path.exists(label_script):
        raise FileNotFoundError(f"CheXbert label.py not found at: {label_script}")

    os.makedirs(output_dir, exist_ok=True)

    cmd = [
        sys.executable, label_script,
        "-d", csv_path,
        "-o", output_dir,
        "-c", checkpoint_path,
    ]

    print(f"\n[INFO] Running CheXbert labeling...")
    print(f"[INFO] Command: {' '.join(cmd)}")

    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        cwd=os.path.dirname(os.path.abspath(__file__)),
    )

    if result.returncode != 0:
        print(f"[ERROR] CheXbert labeling failed!")
        print(f"[STDERR] {result.stderr}")
        raise RuntimeError(f"CheXbert label.py failed with return code {result.returncode}")

    print(result.stdout)

    output_csv = os.path.join(output_dir, "labeled_reports.csv")
    if not os.path.exists(output_csv):
        raise FileNotFoundError(f"Expected output not found: {output_csv}")

    print(f"[INFO] CheXbert labeling complete: {output_csv}")
    return output_csv


# ============================================================
# Step 4: Read labeled results and convert to binary
# ============================================================
def load_labeled_csv(labeled_csv_path):
    """Load CheXbert labeled_reports.csv and convert to binary labels."""
    df = pd.read_csv(labeled_csv_path)
    label_cols = [c for c in df.columns if c != "Report Impression"]
    labels_raw = df[label_cols].copy()
    labels_binary = np.where(labels_raw.values == 1.0, 1.0, 0.0)

    print(f"[INFO] Loaded labeled CSV: {labeled_csv_path}")
    print(f"[INFO] Shape: {labels_binary.shape}, Columns: {label_cols}")

    return labels_binary, labels_raw, label_cols


# ============================================================
# Step 5: Compute Metrics
# ============================================================
def compute_all_metrics(y_true, y_pred, label_names):
    """Compute per-class and macro-averaged metrics."""
    results = {}
    valid_aucs = []
    valid_f1s = []
    valid_recalls = []
    valid_specificities = []

    for i, name in enumerate(label_names):
        yt = y_true[:, i].astype(int)
        yp = y_pred[:, i].astype(int)

        n_pos = int((yt == 1).sum())
        n_neg = int((yt == 0).sum())

        unique = np.unique(yt)
        if len(unique) < 2:
            results[name] = {
                "auc": None, "f1": None, "recall": None, "specificity": None,
                "n_positive": n_pos, "n_negative": n_neg,
                "note": "Skipped: only one class present in ground truth",
            }
            continue

        try:
            auc = roc_auc_score(yt, yp)
        except ValueError:
            auc = None

        f1 = f1_score(yt, yp, zero_division=0)
        recall = recall_score(yt, yp, zero_division=0)

        tn, fp, fn, tp = confusion_matrix(yt, yp, labels=[0, 1]).ravel()
        specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0

        results[name] = {
            "auc": round(auc, 4) if auc is not None else None,
            "f1": round(f1, 4),
            "recall": round(recall, 4),
            "specificity": round(specificity, 4),
            "n_positive": n_pos,
            "n_negative": n_neg,
        }

        if auc is not None:
            valid_aucs.append(auc)
        valid_f1s.append(f1)
        valid_recalls.append(recall)
        valid_specificities.append(specificity)

    results["__macro_avg__"] = {
        "auc": round(np.mean(valid_aucs), 4) if valid_aucs else None,
        "f1": round(np.mean(valid_f1s), 4) if valid_f1s else None,
        "recall": round(np.mean(valid_recalls), 4) if valid_recalls else None,
        "specificity": round(np.mean(valid_specificities), 4) if valid_specificities else None,
    }

    return results


def print_results(results, label_names, title="Results"):
    """Print results in a formatted table."""
    print(f"\n{'=' * 90}")
    print(f"  {title}")
    print(f"{'=' * 90}")
    header = f"{'Label':<35} {'AUC':>8} {'F1':>8} {'Recall':>8} {'Spec':>8} {'N_pos':>8} {'N_neg':>8}"
    print(header)
    print("-" * 90)

    for name in label_names:
        r = results.get(name, {})
        auc = r.get("auc")
        f1 = r.get("f1")
        recall_val = r.get("recall")
        spec = r.get("specificity")
        n_pos = r.get("n_positive", 0)
        n_neg = r.get("n_negative", 0)

        auc_str = f"{auc:.4f}" if auc is not None else "N/A"
        f1_str = f"{f1:.4f}" if f1 is not None else "N/A"
        recall_str = f"{recall_val:.4f}" if recall_val is not None else "N/A"
        spec_str = f"{spec:.4f}" if spec is not None else "N/A"

        print(f"{name:<35} {auc_str:>8} {f1_str:>8} {recall_str:>8} {spec_str:>8} {n_pos:>8} {n_neg:>8}")

    print("-" * 90)
    macro = results.get("__macro_avg__", {})
    m_auc = macro.get("auc")
    m_f1 = macro.get("f1")
    m_recall = macro.get("recall")
    m_spec = macro.get("specificity")
    m_auc_str = f"{m_auc:.4f}" if m_auc is not None else "N/A"
    m_f1_str = f"{m_f1:.4f}" if m_f1 is not None else "N/A"
    m_recall_str = f"{m_recall:.4f}" if m_recall is not None else "N/A"
    m_spec_str = f"{m_spec:.4f}" if m_spec is not None else "N/A"
    print(f"{'MACRO AVERAGE':<35} {m_auc_str:>8} {m_f1_str:>8} {m_recall_str:>8} {m_spec_str:>8}")
    print(f"{'=' * 90}")


# ============================================================
# Save results to xlsx
# ============================================================
def save_results_to_xlsx(results, label_names, args, llm_samples, xlsx_path):
    """Save evaluation results to an xlsx file."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center_align = Alignment(horizontal="center", vertical="center")
    label_font = Font(bold=True, size=10)
    macro_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Sheet 1: Per-Disease Metrics
    ws = wb.active
    ws.title = "Per-Disease Metrics"

    headers = ["Disease", "AUC", "F1", "Recall", "Specificity", "N_Positive", "N_Negative"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    row_idx = 2
    for name in label_names:
        r = results.get(name, {})
        ws.cell(row=row_idx, column=1, value=name).font = label_font
        ws.cell(row=row_idx, column=2, value=r.get("auc"))
        ws.cell(row=row_idx, column=3, value=r.get("f1"))
        ws.cell(row=row_idx, column=4, value=r.get("recall"))
        ws.cell(row=row_idx, column=5, value=r.get("specificity"))
        ws.cell(row=row_idx, column=6, value=r.get("n_positive"))
        ws.cell(row=row_idx, column=7, value=r.get("n_negative"))
        for ci in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=ci).border = border
            if ci > 1:
                ws.cell(row=row_idx, column=ci).alignment = center_align
                ws.cell(row=row_idx, column=ci).number_format = '0.0000'
        row_idx += 1

    macro = results.get("__macro_avg__", {})
    ws.cell(row=row_idx, column=1, value="MACRO AVERAGE").font = Font(bold=True, size=11)
    ws.cell(row=row_idx, column=2, value=macro.get("auc"))
    ws.cell(row=row_idx, column=3, value=macro.get("f1"))
    ws.cell(row=row_idx, column=4, value=macro.get("recall"))
    ws.cell(row=row_idx, column=5, value=macro.get("specificity"))
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=ci)
        cell.border = border
        cell.fill = macro_fill
        if ci > 1:
            cell.alignment = center_align
            cell.number_format = '0.0000'

    col_widths = [30, 10, 10, 10, 12, 12, 12]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Sheet 2: Run Info
    ws2 = wb.create_sheet("Run Info")
    info_items = [
        ("LLM Output", args.llm_output),
        ("CheXbert Checkpoint", args.chexbert_checkpoint),
        ("Label Strategy", "U-zeros (only val==1.0 is positive, everything else is 0.0)"),
        ("Total Reports", len(llm_samples)),
    ]
    ws2.cell(row=1, column=1, value="Parameter").font = Font(bold=True, size=11)
    ws2.cell(row=1, column=2, value="Value").font = Font(bold=True, size=11)
    for ri, (k, v) in enumerate(info_items, start=2):
        ws2.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=str(v))
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 80

    wb.save(xlsx_path)


# ============================================================
# Main
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="Evaluate LLM reports using CheXbert label extraction"
    )
    parser.add_argument(
        "--llm_output", type=str, required=True,
        help="Path to LLM output JSON (e.g., qwen_output.json)",
    )
    parser.add_argument(
        "--annotation_json", type=str, required=True,
        help="Path to annotation.json for ground truth reports",
    )
    parser.add_argument(
        "--chexbert_checkpoint", type=str,
        default="./checkpoints/chexbert.pth",
        help="Path to CheXbert checkpoint (.pth). "
             "Download from: https://stanfordmedicine.app.box.com/s/c3stck6w6dol3h36grdc97xoydzxd7w9",
    )
    parser.add_argument(
        "--output_dir", type=str,
        default="./results/chexbert_eval",
        help="Output directory for results",
    )
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    input_stem = os.path.splitext(os.path.basename(args.llm_output))[0]
    file_prefix = input_stem
    print(f"[INFO] Output file prefix: {file_prefix}")

    # 1. Load LLM output
    llm_samples = load_llm_output(args.llm_output)
    dicom_ids = [s["id"] for s in llm_samples]

    # 2-3. Generate CSVs & run CheXbert
    tmp_dir = tempfile.mkdtemp(prefix="chexbert_tmp_")
    print(f"[INFO] Using temp directory: {tmp_dir}")

    try:
        pred_csv_path = os.path.join(tmp_dir, f"{file_prefix}_pred.csv")
        gt_csv_path = os.path.join(tmp_dir, f"{file_prefix}_gt.csv")

        create_prediction_csv(llm_samples, pred_csv_path)
        create_ground_truth_csv(args.annotation_json, dicom_ids, gt_csv_path)

        pred_label_dir = os.path.join(tmp_dir, "pred_labeled")
        gt_label_dir = os.path.join(tmp_dir, "gt_labeled")

        print(f"\n{'='*60}")
        print(f"  Labeling PREDICTION reports with CheXbert")
        print(f"{'='*60}")
        pred_labeled_csv = run_chexbert_labeling(pred_csv_path, pred_label_dir, args.chexbert_checkpoint)

        print(f"\n{'='*60}")
        print(f"  Labeling GROUND TRUTH reports with CheXbert")
        print(f"{'='*60}")
        gt_labeled_csv = run_chexbert_labeling(gt_csv_path, gt_label_dir, args.chexbert_checkpoint)

        # 4. Load labeled results
        pred_binary, pred_raw, pred_cols = load_labeled_csv(pred_labeled_csv)
        gt_binary, gt_raw, gt_cols = load_labeled_csv(gt_labeled_csv)

        assert pred_binary.shape == gt_binary.shape, \
            f"Shape mismatch: pred {pred_binary.shape} vs gt {gt_binary.shape}"

        label_names = pred_cols

        # 5. Compute Metrics
        print(f"\n[INFO] Evaluating on {pred_binary.shape[0]} samples")
        results = compute_all_metrics(gt_binary, pred_binary, label_names)
        print_results(
            results, label_names,
            title="CheXbert Evaluation: LLM Predictions vs Ground Truth (U-zeros strategy)"
        )

        # 6. Save Results
        xlsx_path = os.path.join(args.output_dir, f"{file_prefix}_eval_summary.xlsx")
        save_results_to_xlsx(results, label_names, args, llm_samples, xlsx_path)
        print(f"\n[INFO] Summary xlsx saved to {xlsx_path}")

        # Save per-sample predictions CSV
        pred_records = []
        for i, did in enumerate(dicom_ids):
            record = {"dicom_id": did}
            for j, label_name in enumerate(label_names):
                raw_pred_val = pred_raw.iloc[i][label_name]
                raw_gt_val = gt_raw.iloc[i][label_name]
                record[f"pred_raw_{label_name}"] = raw_pred_val if not pd.isna(raw_pred_val) else ""
                record[f"gt_raw_{label_name}"] = raw_gt_val if not pd.isna(raw_gt_val) else ""
                record[f"pred_{label_name}"] = int(pred_binary[i, j])
                record[f"gt_{label_name}"] = int(gt_binary[i, j])
            n_agree = sum(
                int(pred_binary[i, j]) == int(gt_binary[i, j])
                for j in range(len(label_names))
            )
            record["n_agree"] = n_agree
            record["n_disagree"] = len(label_names) - n_agree
            pred_records.append(record)

        pred_df = pd.DataFrame(pred_records)
        pred_path = os.path.join(args.output_dir, f"{file_prefix}_per_sample.csv")
        pred_df.to_csv(pred_path, index=False)
        print(f"[INFO] Per-sample predictions saved to {pred_path}")

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        print(f"[INFO] Cleaned up temp directory: {tmp_dir}")

    print(f"\n[DONE] CheXbert evaluation complete!")
    print(f"  Results directory: {args.output_dir}")


if __name__ == "__main__":
    main()
