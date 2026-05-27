"""
LLM-as-Labeler Clinical Accuracy Evaluation
=============================================
Evaluate LLM-generated radiology reports by using ANOTHER LLM (e.g. Qwen3.5)
as a 14-class CheXpert label extractor.

Pipeline:
    report text --> Labeler LLM --> 14-class {0/1} JSON --> metrics

Both the LLM-generated reports (predictions) AND the ground truth reports
are labeled by the same labeler LLM, so the comparison is apples-to-apples.

Usage:
    CUDA_VISIBLE_DEVICES=0,1,2,3 python evaluate_llm_as_labeler.py \
        --llm_output ./results/qwen_output.json \
        --annotation_json ./data/annotation.json \
        --labeler_model Qwen/Qwen2.5-7B-Instruct \
        --output_dir ./results/llm_as_labeler_eval
"""

import os
import sys
import json
import re
import time
import argparse
import warnings

import numpy as np
import pandas as pd
import torch
from tqdm import tqdm
from transformers import AutoModelForCausalLM, AutoTokenizer

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ============================================================
# Constants
# ============================================================

CHEXPERT_LABELS = [
    "Enlarged Cardiomediastinum",
    "Cardiomegaly",
    "Lung Opacity",
    "Lung Lesion",
    "Edema",
    "Consolidation",
    "Pneumonia",
    "Atelectasis",
    "Pneumothorax",
    "Pleural Effusion",
    "Pleural Other",
    "Fracture",
    "Support Devices",
    "No Finding",
]

# ============================================================
# Prompt
# ============================================================

LABELER_SYSTEM_PROMPT = (
    "You are a clinical information-extraction assistant for chest X-ray "
    "radiology reports. Given a report, you must decide, for each of the 14 "
    "CheXpert conditions, whether the report asserts the condition is PRESENT."
)

LABELER_USER_PROMPT_TEMPLATE = (
    "Read the following chest X-ray report and, for each of the 14 CheXpert "
    "conditions listed below, output 1 if the report asserts the condition is "
    "PRESENT (explicit positive finding), otherwise output 0.\n\n"
    "Labeling rules (U-zeros, MUST follow strictly):\n"
    "  - Explicitly denied (e.g. \"no pneumothorax\", \"without effusion\") -> 0\n"
    "  - Uncertain / hedged (e.g. \"possible\", \"cannot exclude\", \"may represent\") -> 0\n"
    "  - Not mentioned in the report -> 0\n"
    "  - Only explicit positive assertions -> 1\n"
    "  - \"No Finding\" = 1 if and only if the report explicitly states the study "
    "is normal / shows no acute abnormality.\n\n"
    "The 14 conditions (use these exact keys, in this exact order):\n"
    "  Enlarged Cardiomediastinum, Cardiomegaly, Lung Opacity, Lung Lesion, "
    "Edema, Consolidation, Pneumonia, Atelectasis, Pneumothorax, "
    "Pleural Effusion, Pleural Other, Fracture, Support Devices, No Finding\n\n"
    "Report:\n"
    "\"\"\"\n"
    "{report_text}\n"
    "\"\"\"\n\n"
    "Respond with STRICT JSON only (no prose, no markdown, no code fences). "
    "The JSON object must contain exactly these 14 keys, each mapped to an "
    "integer 0 or 1. Example format:\n"
    "{{\"Enlarged Cardiomediastinum\": 0, \"Cardiomegaly\": 1, \"Lung Opacity\": 0, "
    "\"Lung Lesion\": 0, \"Edema\": 0, \"Consolidation\": 0, \"Pneumonia\": 0, "
    "\"Atelectasis\": 1, \"Pneumothorax\": 0, \"Pleural Effusion\": 1, "
    "\"Pleural Other\": 0, \"Fracture\": 0, \"Support Devices\": 0, "
    "\"No Finding\": 0}}"
)

# ============================================================
# Metrics (self-contained, no cross-file import)
# ============================================================

from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix


def compute_all_metrics(y_true, y_pred, label_names):
    """Compute per-class and macro-averaged metrics."""
    results = {}
    valid_aucs = []
    valid_f1s = []
    valid_recalls = []
    valid_specificities = []

    for i, name in enumerate(label_names):
        yt = y_true[:, i].astype(int)
        yp = y_pred[:, i].astype(int)

        n_pos = int((yt == 1).sum())
        n_neg = int((yt == 0).sum())

        unique = np.unique(yt)
        if len(unique) < 2:
            results[name] = {
                "auc": None, "f1": None, "recall": None, "specificity": None,
                "n_positive": n_pos, "n_negative": n_neg,
                "note": "Skipped: only one class present in ground truth",
            }
            continue

        try:
            auc = roc_auc_score(yt, yp)
        except ValueError:
            auc = None

        f1 = f1_score(yt, yp, zero_division=0)
        recall = recall_score(yt, yp, zero_division=0)

        tn, fp, fn, tp = confusion_matrix(yt, yp, labels=[0, 1]).ravel()
        specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0

        results[name] = {
            "auc": round(auc, 4) if auc is not None else None,
            "f1": round(f1, 4),
            "recall": round(recall, 4),
            "specificity": round(specificity, 4),
            "n_positive": n_pos,
            "n_negative": n_neg,
        }

        if auc is not None:
            valid_aucs.append(auc)
        valid_f1s.append(f1)
        valid_recalls.append(recall)
        valid_specificities.append(specificity)

    results["__macro_avg__"] = {
        "auc": round(np.mean(valid_aucs), 4) if valid_aucs else None,
        "f1": round(np.mean(valid_f1s), 4) if valid_f1s else None,
        "recall": round(np.mean(valid_recalls), 4) if valid_recalls else None,
        "specificity": round(np.mean(valid_specificities), 4) if valid_specificities else None,
    }

    return results


def print_results(results, label_names, title="Results"):
    """Print results in a formatted table."""
    print(f"\n{'=' * 90}")
    print(f"  {title}")
    print(f"{'=' * 90}")
    header = f"{'Label':<35} {'AUC':>8} {'F1':>8} {'Recall':>8} {'Spec':>8} {'N_pos':>8} {'N_neg':>8}"
    print(header)
    print("-" * 90)

    for name in label_names:
        r = results.get(name, {})
        auc = r.get("auc")
        f1 = r.get("f1")
        recall_val = r.get("recall")
        spec = r.get("specificity")
        n_pos = r.get("n_positive", 0)
        n_neg = r.get("n_negative", 0)

        auc_str = f"{auc:.4f}" if auc is not None else "N/A"
        f1_str = f"{f1:.4f}" if f1 is not None else "N/A"
        recall_str = f"{recall_val:.4f}" if recall_val is not None else "N/A"
        spec_str = f"{spec:.4f}" if spec is not None else "N/A"

        print(f"{name:<35} {auc_str:>8} {f1_str:>8} {recall_str:>8} {spec_str:>8} {n_pos:>8} {n_neg:>8}")

    print("-" * 90)
    macro = results.get("__macro_avg__", {})
    m_auc = macro.get("auc")
    m_f1 = macro.get("f1")
    m_recall = macro.get("recall")
    m_spec = macro.get("specificity")
    print(f"{'MACRO AVERAGE':<35} "
          f"{m_auc:.4f if m_auc else 'N/A':>8} "
          f"{m_f1:.4f if m_f1 else 'N/A':>8} "
          f"{m_recall:.4f if m_recall else 'N/A':>8} "
          f"{m_spec:.4f if m_spec else 'N/A':>8}")
    print(f"{'=' * 90}")


# ============================================================
# Data loading
# ============================================================

def load_llm_output(json_path):
    """Load LLM-generated reports from JSON file."""
    print(f"[INFO] Loading LLM output from: {json_path}")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    samples = data.get("test", data if isinstance(data, list) else [])
    print(f"[INFO] Loaded {len(samples)} LLM-generated reports")
    return samples


def load_ground_truth_reports(annotation_json_path, dicom_ids):
    """Extract ground truth report text for the given dicom_ids."""
    print(f"[INFO] Loading annotation.json for ground truth reports...")
    with open(annotation_json_path, "r", encoding="utf-8") as f:
        annotation = json.load(f)

    dicom_to_report = {}
    for split_name in ["train", "val", "test"]:
        if split_name not in annotation:
            continue
        for sample in annotation[split_name]:
            did = sample["id"]
            dicom_to_report[did] = sample.get("report", "")

    reports = []
    n_missing = 0
    for did in dicom_ids:
        report = dicom_to_report.get(did, "")
        if not report:
            n_missing += 1
        reports.append(report)
    if n_missing > 0:
        print(f"[WARN] {n_missing}/{len(dicom_ids)} dicom_ids had no report in annotation.json")
    return reports


def clean_report(text):
    """Strip LLM artifacts from report text."""
    if not text:
        return ""
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL)
    text = re.sub(r"<\|im_end\|>", "", text)
    text = re.sub(r"<\|im_start\|>", "", text)
    text = re.sub(r"\*\*([^*]+)\*\*", r"\1", text)
    text = re.sub(r"^#+\s+", "", text, flags=re.MULTILINE)
    text = re.sub(r"^\s*[-*•]\s+", "", text, flags=re.MULTILINE)
    text = " ".join(text.split())
    return text.strip()


# ============================================================
# LLM output parsing
# ============================================================

def _extract_json_blob(text: str) -> str:
    """Extract the first balanced {...} JSON object from text."""
    if not text:
        return ""
    text = re.sub(r"```(?:json)?", "", text, flags=re.IGNORECASE)
    text = text.replace("```", "")
    start = text.find("{")
    if start == -1:
        return ""
    depth = 0
    for i in range(start, len(text)):
        c = text[i]
        if c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return ""


def parse_labeler_output(raw_text: str):
    """Parse the labeler LLM's raw output into a 14-dim binary vector."""
    if not raw_text:
        return np.zeros(len(CHEXPERT_LABELS), dtype=int), False, "empty_output"

    blob = _extract_json_blob(raw_text)
    if not blob:
        return np.zeros(len(CHEXPERT_LABELS), dtype=int), False, "no_json_blob"

    try:
        obj = json.loads(blob)
    except json.JSONDecodeError as e:
        return np.zeros(len(CHEXPERT_LABELS), dtype=int), False, f"json_decode_error: {e}"

    if not isinstance(obj, dict):
        return np.zeros(len(CHEXPERT_LABELS), dtype=int), False, "not_a_dict"

    lower_map = {k.lower().strip(): v for k, v in obj.items()}

    vec = np.zeros(len(CHEXPERT_LABELS), dtype=int)
    missing = []
    for i, name in enumerate(CHEXPERT_LABELS):
        key = name.lower().strip()
        if key not in lower_map:
            missing.append(name)
            continue
        v = lower_map[key]
        try:
            iv = int(v)
        except (TypeError, ValueError):
            if isinstance(v, str):
                vs = v.strip().lower()
                if vs in ("1", "true", "yes", "positive", "present"):
                    iv = 1
                else:
                    iv = 0
            else:
                iv = 0
        vec[i] = 1 if iv == 1 else 0

    if missing:
        return vec, False, f"missing_keys: {missing}"
    return vec, True, ""


# ============================================================
# LLM labeler
# ============================================================

def load_labeler_model(model_name: str, flash_attn: bool = False):
    """Load the labeler LLM (text-only, AutoModelForCausalLM)."""
    print(f"[INFO] Loading labeler LLM: {model_name}")
    model_kwargs = {
        "torch_dtype": torch.bfloat16,
        "device_map": "auto",
    }
    if flash_attn:
        model_kwargs["attn_implementation"] = "flash_attention_2"
        print(f"[INFO] Using flash_attention_2")

    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = AutoModelForCausalLM.from_pretrained(model_name, **model_kwargs)
    model.eval()

    if torch.cuda.is_available():
        n_gpus = torch.cuda.device_count()
        print(f"[INFO] Labeler loaded across {n_gpus} GPU(s)")

    return model, tokenizer


@torch.no_grad()
def run_labeler_on_reports(
    reports, model, tokenizer, max_new_tokens=256,
    desc="Labeling", resume_raw=None,
):
    """Run the labeler LLM on a list of report strings."""
    n = len(reports)
    labels_binary = np.zeros((n, len(CHEXPERT_LABELS)), dtype=int)
    raw_outputs = [""] * n
    parse_errors = [""] * n

    for i in tqdm(range(n), desc=desc):
        reused = False
        if resume_raw is not None and i < len(resume_raw):
            prev = resume_raw[i]
            if isinstance(prev, str) and prev.strip():
                raw_outputs[i] = prev
                reused = True

        if not reused:
            report_text = reports[i] if reports[i] else "(empty report)"
            if len(report_text) > 8000:
                report_text = report_text[:8000]

            user_msg = LABELER_USER_PROMPT_TEMPLATE.format(report_text=report_text)
            messages = [
                {"role": "system", "content": LABELER_SYSTEM_PROMPT},
                {"role": "user",   "content": user_msg},
            ]
            try:
                prompt_text = tokenizer.apply_chat_template(
                    messages,
                    tokenize=False,
                    add_generation_prompt=True,
                    enable_thinking=False,
                )
            except TypeError:
                prompt_text = tokenizer.apply_chat_template(
                    messages,
                    tokenize=False,
                    add_generation_prompt=True,
                )

            inputs = tokenizer(prompt_text, return_tensors="pt").to(model.device)
            try:
                generated = model.generate(
                    **inputs,
                    max_new_tokens=max_new_tokens,
                    do_sample=False,
                    temperature=1.0,
                    pad_token_id=tokenizer.eos_token_id,
                )
                gen_trimmed = generated[0][inputs["input_ids"].shape[-1]:]
                out_text = tokenizer.decode(gen_trimmed, skip_special_tokens=True)
            except Exception as e:
                out_text = ""
                parse_errors[i] = f"generate_error: {str(e)[:200]}"

            raw_outputs[i] = out_text

        vec, ok, err = parse_labeler_output(raw_outputs[i])
        labels_binary[i] = vec
        if not ok and not parse_errors[i]:
            parse_errors[i] = err

    n_parse_fail = sum(1 for e in parse_errors if e)
    print(f"[INFO] {desc}: parse failures = {n_parse_fail}/{n}")

    return labels_binary, raw_outputs, parse_errors


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Evaluate LLM-generated reports using an LLM as the "
                    "14-class CheXpert label extractor."
    )
    parser.add_argument(
        "--llm_output", type=str, required=True,
        help="Path to LLM-generated reports JSON.",
    )
    parser.add_argument(
        "--annotation_json", type=str, required=True,
        help="Path to annotation.json for ground truth reports.",
    )
    parser.add_argument(
        "--labeler_model", type=str,
        default="Qwen/Qwen2.5-7B-Instruct",
        help="HuggingFace model name or local path of the labeler LLM (text-only).",
    )
    parser.add_argument(
        "--output_dir", type=str,
        default="./results/llm_as_labeler_eval",
        help="Output directory for results.",
    )
    parser.add_argument(
        "--max_new_tokens", type=int, default=256,
        help="Max new tokens for the labeler output.",
    )
    parser.add_argument(
        "--flash_attn", action="store_true", default=False,
        help="Use flash_attention_2 for faster inference.",
    )
    parser.add_argument(
        "--resume", action="store_true", default=False,
        help="Resume from existing labeler_raw.json.",
    )
    parser.add_argument(
        "--gt_cache_dir", type=str, default=None,
        help="Directory for persistent GT-labels cache (default: --output_dir).",
    )
    parser.add_argument(
        "--max_samples", type=int, default=None,
        help="Max samples to process (for debugging).",
    )
    parser.add_argument(
        "--no_clean_reports", action="store_true", default=False,
        help="Do NOT clean report text before labeling.",
    )
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    file_prefix = os.path.splitext(os.path.basename(args.llm_output))[0]
    print(f"[INFO] Output file prefix: {file_prefix}")

    raw_json_path = os.path.join(args.output_dir, f"{file_prefix}_labeler_raw.json")
    xlsx_path = os.path.join(args.output_dir, f"{file_prefix}_eval.xlsx")

    # GT cache
    gt_cache_dir = args.gt_cache_dir if args.gt_cache_dir else args.output_dir
    os.makedirs(gt_cache_dir, exist_ok=True)
    labeler_tag = os.path.basename(os.path.normpath(args.labeler_model))
    labeler_tag = re.sub(r"[^A-Za-z0-9._-]+", "_", labeler_tag) or "labeler"
    gt_cache_path = os.path.join(gt_cache_dir, f"gt_labels_cache__{labeler_tag}.json")
    print(f"[INFO] GT labels cache file: {gt_cache_path}")

    # 1. Load reports
    llm_samples = load_llm_output(args.llm_output)
    if args.max_samples is not None:
        llm_samples = llm_samples[:args.max_samples]

    dicom_ids = [s["id"] for s in llm_samples]
    pred_reports_raw = [s.get("output", "") for s in llm_samples]
    gt_reports_raw = load_ground_truth_reports(args.annotation_json, dicom_ids)

    if args.no_clean_reports:
        pred_reports = pred_reports_raw
        gt_reports = gt_reports_raw
    else:
        print(f"[INFO] Cleaning reports...")
        pred_reports = [clean_report(r) for r in pred_reports_raw]
        gt_reports = [clean_report(r) for r in gt_reports_raw]

    n_total = len(dicom_ids)

    # 2. Resume support
    prev_pred_raw = None
    prev_gt_raw = None
    if args.resume and os.path.exists(raw_json_path):
        print(f"[INFO] Resuming from {raw_json_path}")
        with open(raw_json_path, "r", encoding="utf-8") as f:
            prev = json.load(f)
        prev_pred_raw = prev.get("pred_raw_outputs", None)
        prev_gt_raw = prev.get("gt_raw_outputs", None)

    # GT cache
    gt_cache_entries = {}
    if os.path.exists(gt_cache_path):
        try:
            with open(gt_cache_path, "r", encoding="utf-8") as f:
                _gc = json.load(f)
            gt_cache_entries = _gc.get("entries", {}) or {}
            print(f"[INFO] Loaded GT cache: {len(gt_cache_entries)} entries")
        except Exception:
            gt_cache_entries = {}

    gt_resume_from_cache = [
        (gt_cache_entries.get(did, {}) or {}).get("raw", "") for did in dicom_ids
    ]
    n_cache_hits = sum(1 for r in gt_resume_from_cache if isinstance(r, str) and r.strip())
    print(f"[INFO] GT cache hits: {n_cache_hits}/{n_total}")

    if prev_gt_raw is not None:
        for i in range(min(len(prev_gt_raw), len(gt_resume_from_cache))):
            if isinstance(prev_gt_raw[i], str) and prev_gt_raw[i].strip():
                gt_resume_from_cache[i] = prev_gt_raw[i]

    # 3. Load labeler LLM
    model, tokenizer = load_labeler_model(args.labeler_model, flash_attn=args.flash_attn)

    # 4. Label predictions + ground truth
    print(f"\n{'='*60}\n  Labeling PREDICTION reports\n{'='*60}")
    t0 = time.time()
    pred_binary, pred_raw, pred_errs = run_labeler_on_reports(
        pred_reports, model, tokenizer,
        max_new_tokens=args.max_new_tokens,
        desc="Label PRED",
        resume_raw=prev_pred_raw,
    )
    t_pred = time.time() - t0

    print(f"\n{'='*60}\n  Labeling GROUND TRUTH reports\n{'='*60}")
    t0 = time.time()
    gt_binary, gt_raw, gt_errs = run_labeler_on_reports(
        gt_reports, model, tokenizer,
        max_new_tokens=args.max_new_tokens,
        desc="Label GT",
        resume_raw=gt_resume_from_cache,
    )
    t_gt = time.time() - t0

    # Update GT cache
    n_cache_added = 0
    for i, did in enumerate(dicom_ids):
        if not (isinstance(gt_raw[i], str) and gt_raw[i].strip()):
            continue
        if did not in gt_cache_entries:
            n_cache_added += 1
        gt_cache_entries[did] = {
            "raw": gt_raw[i],
            "labels": [int(x) for x in gt_binary[i].tolist()],
            "parse_error": gt_errs[i],
        }
    try:
        with open(gt_cache_path, "w", encoding="utf-8") as f:
            json.dump({
                "labeler_model": args.labeler_model,
                "chexpert_labels": CHEXPERT_LABELS,
                "n_entries": len(gt_cache_entries),
                "entries": gt_cache_entries,
            }, f, indent=2, ensure_ascii=False)
        print(f"[INFO] GT cache updated: +{n_cache_added} new, total {len(gt_cache_entries)}")
    except Exception as e:
        print(f"[WARN] Failed to write GT cache: {e}")

    # 5. Save raw outputs
    with open(raw_json_path, "w", encoding="utf-8") as f:
        json.dump({
            "labeler_model": args.labeler_model,
            "llm_output": args.llm_output,
            "n_samples": n_total,
            "dicom_ids": dicom_ids,
            "pred_raw_outputs": pred_raw,
            "gt_raw_outputs": gt_raw,
            "pred_parse_errors": pred_errs,
            "gt_parse_errors": gt_errs,
        }, f, indent=2, ensure_ascii=False)
    print(f"[INFO] Raw outputs saved to {raw_json_path}")

    # 6. Compute metrics
    results = compute_all_metrics(gt_binary, pred_binary, CHEXPERT_LABELS)
    print_results(results, CHEXPERT_LABELS,
                  title="LLM-as-Labeler Evaluation (U-zeros strategy)")

    # 7. Save xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrics"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["Disease", "AUC", "F1", "Recall", "Specificity", "N_Positive", "N_Negative"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    row_idx = 2
    for name in CHEXPERT_LABELS:
        r = results.get(name, {})
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=r.get("auc"))
        ws.cell(row=row_idx, column=3, value=r.get("f1"))
        ws.cell(row=row_idx, column=4, value=r.get("recall"))
        ws.cell(row=row_idx, column=5, value=r.get("specificity"))
        ws.cell(row=row_idx, column=6, value=r.get("n_positive"))
        ws.cell(row=row_idx, column=7, value=r.get("n_negative"))
        for ci in range(1, 8):
            ws.cell(row=row_idx, column=ci).border = border
        row_idx += 1

    macro = results.get("__macro_avg__", {})
    ws.cell(row=row_idx, column=1, value="MACRO AVERAGE").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value=macro.get("auc"))
    ws.cell(row=row_idx, column=3, value=macro.get("f1"))
    ws.cell(row=row_idx, column=4, value=macro.get("recall"))
    ws.cell(row=row_idx, column=5, value=macro.get("specificity"))

    wb.save(xlsx_path)
    print(f"[INFO] Eval xlsx saved to {xlsx_path}")

    print(f"\n[DONE] LLM-as-Labeler evaluation complete!")
    print(f"  Labeler timing: PRED={t_pred:.1f}s, GT={t_gt:.1f}s")


if __name__ == "__main__":
    main()
