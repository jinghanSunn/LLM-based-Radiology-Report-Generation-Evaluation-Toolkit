"""
NLG + Clinical Report-Level Metrics Evaluation
================================================
Evaluate LLM-generated radiology reports using multiple metrics:

  ① NLG metrics:      BLEU-1/2/3/4, ROUGE-L, METEOR
  ② BERTScore:        Contextual embedding similarity (F1)
  ③ RadGraph F1:      Entity & relation level F1 via RadGraph
  ④ RaTEScore:        Radiology Text Evaluation Score

Input:
  - LLM output JSON: {"test": [{"id": "dicom_id", "output": "report text"}, ...]}
  - annotation.json: ground truth reports

Output:
  - <prefix>_nlg_metrics.xlsx: summary metrics
  - <prefix>_nlg_per_sample.csv: per-sample scores
  - <prefix>_nlg_metrics.json: machine-readable metrics

Usage:
    python evaluate_nlg.py \
        --llm_output ./results/qwen_output.json \
        --annotation_json ./data/annotation.json \
        --output_dir ./results/nlg_eval \
        --metrics bleu,rouge,meteor,bertscore
"""

import os
import sys
import json
import re
import argparse
import warnings
import time
from collections import defaultdict

import numpy as np
import pandas as pd
from tqdm import tqdm

warnings.filterwarnings("ignore")


# ============================================================
# Metric Implementations
# ============================================================

def compute_bleu(references, hypotheses):
    """Compute corpus-level BLEU-1/2/3/4 using NLTK."""
    from nltk.translate.bleu_score import corpus_bleu, sentence_bleu, SmoothingFunction

    smooth = SmoothingFunction().method1

    refs_tokenized = [[ref.lower().split()] for ref in references]
    hyps_tokenized = [hyp.lower().split() for hyp in hypotheses]

    results = {}
    for n in range(1, 5):
        weights = tuple([1.0 / n] * n + [0.0] * (4 - n))
        score = corpus_bleu(refs_tokenized, hyps_tokenized, weights=weights,
                            smoothing_function=smooth)
        results[f"BLEU-{n}"] = round(score, 4)

    per_sample = []
    for ref_tok, hyp_tok in tqdm(zip(refs_tokenized, hyps_tokenized),
                                  total=len(refs_tokenized), desc="BLEU"):
        sample_scores = {}
        for n in range(1, 5):
            weights = tuple([1.0 / n] * n + [0.0] * (4 - n))
            s = sentence_bleu(ref_tok, hyp_tok, weights=weights,
                              smoothing_function=smooth)
            sample_scores[f"BLEU-{n}"] = round(s, 4)
        per_sample.append(sample_scores)

    return results, per_sample


def compute_rouge(references, hypotheses):
    """Compute ROUGE-L using google rouge-score library."""
    from rouge_score import rouge_scorer

    scorer = rouge_scorer.RougeScorer(['rougeL'], use_stemmer=True)

    per_sample = []
    all_scores = []
    for ref, hyp in tqdm(zip(references, hypotheses),
                          total=len(references), desc="ROUGE-L"):
        scores = scorer.score(ref, hyp)
        f1 = scores['rougeL'].fmeasure
        all_scores.append(f1)
        per_sample.append({"ROUGE-L": round(f1, 4)})

    results = {"ROUGE-L": round(np.mean(all_scores), 4)}
    return results, per_sample


def compute_meteor(references, hypotheses):
    """Compute METEOR using NLTK."""
    from nltk.translate.meteor_score import meteor_score
    import nltk
    for resource, name in [('corpora/wordnet', 'wordnet'), ('corpora/omw-1.4', 'omw-1.4')]:
        try:
            nltk.data.find(resource)
        except (LookupError, Exception):
            nltk.download(name, quiet=True)

    per_sample = []
    all_scores = []
    for ref, hyp in tqdm(zip(references, hypotheses),
                          total=len(references), desc="METEOR"):
        ref_tokens = ref.lower().split()
        hyp_tokens = hyp.lower().split()
        if len(hyp_tokens) == 0:
            score = 0.0
        else:
            score = meteor_score([ref_tokens], hyp_tokens)
        all_scores.append(score)
        per_sample.append({"METEOR": round(score, 4)})

    results = {"METEOR": round(np.mean(all_scores), 4)}
    return results, per_sample


def compute_bertscore(references, hypotheses, batch_size=64, model_type=None):
    """Compute BERTScore (Precision, Recall, F1)."""
    from bert_score import score as bert_score_fn

    model_name = model_type or "roberta-large"
    print(f"[INFO] Computing BERTScore with model: {model_name} ...")

    score_kwargs = dict(
        cands=hypotheses,
        refs=references,
        verbose=True,
        batch_size=batch_size,
    )
    if model_type:
        score_kwargs["model_type"] = model_type
        score_kwargs["rescale_with_baseline"] = False
    else:
        score_kwargs["lang"] = "en"
        score_kwargs["rescale_with_baseline"] = True

    P, R, F1 = bert_score_fn(**score_kwargs)

    P = P.numpy()
    R = R.numpy()
    F1 = F1.numpy()

    per_sample = []
    for p, r, f in zip(P, R, F1):
        per_sample.append({
            "BERTScore-P": round(float(p), 4),
            "BERTScore-R": round(float(r), 4),
            "BERTScore-F1": round(float(f), 4),
        })

    results = {
        "BERTScore-P": round(float(np.mean(P)), 4),
        "BERTScore-R": round(float(np.mean(R)), 4),
        "BERTScore-F1": round(float(np.mean(F1)), 4),
    }
    return results, per_sample


def compute_radgraph_f1(references, hypotheses, model_type="modern-radgraph-xl"):
    """Compute RadGraph F1 (entity-level)."""
    try:
        from radgraph import RadGraph
    except ImportError:
        print("[WARN] radgraph not installed. Skipping RadGraph F1.")
        print("[WARN] Install with: pip install radgraph")
        return None, None

    print(f"[INFO] Loading RadGraph model: {model_type} ...")
    try:
        radgraph = RadGraph(model_type=model_type)
    except Exception as e:
        print(f"[ERROR] Failed to load RadGraph model: {e}")
        return None, None

    def _entity_f1(ann_hyp, ann_ref):
        def extract_entities(ann):
            entities = set()
            if isinstance(ann, dict) and "entities" in ann:
                for eid, ent in ann["entities"].items():
                    token = ent.get("tokens", "").lower().strip()
                    label = ent.get("label", "").lower().strip()
                    if token and label:
                        entities.add((token, label))
            return entities

        hyp_ents = extract_entities(ann_hyp)
        ref_ents = extract_entities(ann_ref)

        if len(ref_ents) == 0 and len(hyp_ents) == 0:
            return 1.0
        if len(ref_ents) == 0 or len(hyp_ents) == 0:
            return 0.0

        tp = len(hyp_ents & ref_ents)
        precision = tp / len(hyp_ents) if len(hyp_ents) > 0 else 0.0
        recall = tp / len(ref_ents) if len(ref_ents) > 0 else 0.0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
        return f1

    print("[INFO] Annotating hypothesis reports with RadGraph...")
    try:
        hyp_annotations = radgraph(hypotheses)
    except Exception as e:
        print(f"[ERROR] RadGraph annotation of hypotheses failed: {e}")
        return None, None

    print("[INFO] Annotating reference reports with RadGraph...")
    try:
        ref_annotations = radgraph(references)
    except Exception as e:
        print(f"[ERROR] RadGraph annotation of references failed: {e}")
        return None, None

    per_sample = []
    all_f1 = []
    for i in tqdm(range(len(hypotheses)), desc="RadGraph F1"):
        hyp_ann = hyp_annotations.get(str(i), hyp_annotations.get(i, {}))
        ref_ann = ref_annotations.get(str(i), ref_annotations.get(i, {}))
        f1 = _entity_f1(hyp_ann, ref_ann)
        all_f1.append(f1)
        per_sample.append({"RadGraph-F1": round(f1, 4)})

    results = {"RadGraph-F1": round(float(np.mean(all_f1)), 4)}
    return results, per_sample


def compute_ratescore(references, hypotheses, batch_size=8):
    """Compute RaTEScore (Radiology Text Evaluation Score)."""
    try:
        from RaTEScore import RaTEScore
    except ImportError:
        print("[WARN] RaTEScore not installed. Skipping RaTEScore.")
        print("[WARN] Install with: pip install ratescore")
        return None, None

    print("[INFO] Computing RaTEScore...")
    try:
        import io
        import contextlib
        import logging

        for logger_name in ["medspacy", "spacy", "PyRuSH", "PyRuSH.PyRuSHSentencizer"]:
            logging.getLogger(logger_name).setLevel(logging.CRITICAL)

        try:
            from loguru import logger as loguru_logger
            loguru_logger.disable("PyRuSH")
            loguru_logger.remove()
            loguru_logger.add(sys.stderr, level="WARNING")
        except ImportError:
            pass

        scorer = RaTEScore(batch_size=batch_size)

        chunk_size = 50
        all_scores = []
        n_total = len(hypotheses)

        for start_idx in tqdm(range(0, n_total, chunk_size),
                              desc="RaTEScore", total=(n_total + chunk_size - 1) // chunk_size):
            end_idx = min(start_idx + chunk_size, n_total)
            hyp_chunk = hypotheses[start_idx:end_idx]
            ref_chunk = references[start_idx:end_idx]

            with contextlib.redirect_stdout(io.StringIO()), \
                 contextlib.redirect_stderr(io.StringIO()):
                chunk_scores = scorer.compute_score(hyp_chunk, ref_chunk)

            if isinstance(chunk_scores, (list, np.ndarray)):
                all_scores.extend([float(s) for s in chunk_scores])
            else:
                all_scores.extend([float(chunk_scores)] * len(hyp_chunk))

        per_sample = [{"RaTEScore": round(s, 4)} for s in all_scores]
        results = {"RaTEScore": round(float(np.mean(all_scores)), 4)}
        return results, per_sample

    except Exception as e:
        print(f"[ERROR] RaTEScore computation failed: {e}")
        import traceback
        traceback.print_exc()
        return None, None


# ============================================================
# Data Loading
# ============================================================

def load_llm_output(json_path):
    """Load LLM-generated reports from JSON file."""
    print(f"[INFO] Loading LLM output from: {json_path}")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    samples = data.get("test", data if isinstance(data, list) else [])
    print(f"[INFO] Loaded {len(samples)} LLM-generated reports")
    return samples


def load_ground_truth(annotation_json_path, dicom_ids):
    """Load ground truth reports from annotation.json."""
    print(f"[INFO] Loading annotation.json for ground truth reports...")
    with open(annotation_json_path, "r", encoding="utf-8") as f:
        annotation = json.load(f)

    dicom_to_report = {}
    for split_name in ["train", "val", "test"]:
        if split_name not in annotation:
            continue
        for sample in annotation[split_name]:
            did = sample["id"]
            dicom_to_report[did] = sample.get("report", "")

    reports = []
    n_missing = 0
    for did in dicom_ids:
        report = dicom_to_report.get(did, "")
        if not report:
            n_missing += 1
        reports.append(report)

    if n_missing > 0:
        print(f"[WARN] {n_missing}/{len(dicom_ids)} dicom_ids had no report in annotation.json")
    return reports


def clean_report(text):
    """Clean report text for evaluation."""
    if not text:
        return ""
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL)
    text = re.sub(r'<\|im_end\|>', '', text)
    text = re.sub(r'<\|im_start\|>', '', text)
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'^#+\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'^\s*[-*•]\s+', '', text, flags=re.MULTILINE)
    text = ' '.join(text.split())
    return text.strip()


# ============================================================
# Save Results
# ============================================================

def save_results_to_xlsx(corpus_results, args, n_samples, xlsx_path):
    """Save evaluation results to a formatted xlsx file."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center_align = Alignment(horizontal="center", vertical="center")

    ws = wb.active
    ws.title = "Summary Metrics"

    headers = ["Metric", "Value"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    row_idx = 2
    categories = {
        "NLG Metrics": ["BLEU-1", "BLEU-2", "BLEU-3", "BLEU-4", "ROUGE-L", "METEOR"],
        "Embedding Metrics": ["BERTScore-P", "BERTScore-R", "BERTScore-F1"],
        "Clinical Metrics": ["RadGraph-F1", "RaTEScore"],
    }

    for cat_name, metric_keys in categories.items():
        cell = ws.cell(row=row_idx, column=1, value=cat_name)
        cell.font = Font(bold=True, size=11, color="2F5496")
        cell.border = border
        ws.cell(row=row_idx, column=2).border = border
        row_idx += 1

        for key in metric_keys:
            if key in corpus_results:
                ws.cell(row=row_idx, column=1, value=key).border = border
                cell = ws.cell(row=row_idx, column=2, value=corpus_results[key])
                cell.border = border
                cell.alignment = center_align
                cell.number_format = '0.0000'
                row_idx += 1

    row_idx += 1
    info_items = [
        ("LLM Output", args.llm_output),
        ("Total Samples", n_samples),
        ("Metrics Computed", args.metrics),
    ]
    for k, v in info_items:
        ws.cell(row=row_idx, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=str(v))
        row_idx += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60

    wb.save(xlsx_path)
    print(f"[INFO] Summary xlsx saved to {xlsx_path}")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Evaluate LLM reports using NLG + clinical metrics"
    )
    parser.add_argument(
        "--llm_output", type=str, required=True,
        help="Path to LLM output JSON",
    )
    parser.add_argument(
        "--annotation_json", type=str, required=True,
        help="Path to annotation.json for ground truth reports",
    )
    parser.add_argument(
        "--output_dir", type=str,
        default="./results/nlg_eval",
        help="Output directory for results",
    )
    parser.add_argument(
        "--metrics", type=str,
        default="bleu,rouge,meteor,bertscore",
        help="Comma-separated list of metrics to compute. "
             "Options: bleu, rouge, meteor, bertscore, radgraph, ratescore",
    )
    parser.add_argument(
        "--bertscore_batch_size", type=int, default=64,
        help="Batch size for BERTScore computation",
    )
    parser.add_argument(
        "--bertscore_model", type=str, default=None,
        help="BERTScore model (default: roberta-large from HuggingFace). "
             "Can be a local path or HuggingFace model name.",
    )
    parser.add_argument(
        "--radgraph_model", type=str, default="modern-radgraph-xl",
        help="RadGraph model type (default: modern-radgraph-xl)",
    )
    parser.add_argument(
        "--no_clean_reports", action="store_true", default=False,
        help="Do NOT clean report text before evaluation",
    )
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    input_stem = os.path.splitext(os.path.basename(args.llm_output))[0]
    file_prefix = input_stem
    print(f"[INFO] Output file prefix: {file_prefix}")

    metrics_to_compute = [m.strip().lower() for m in args.metrics.split(",")]
    print(f"[INFO] Metrics to compute: {metrics_to_compute}")

    # 1. Load data
    llm_samples = load_llm_output(args.llm_output)
    dicom_ids = [s["id"] for s in llm_samples]
    hypotheses_raw = [s.get("output", "") for s in llm_samples]
    references_raw = load_ground_truth(args.annotation_json, dicom_ids)

    if not args.no_clean_reports:
        print("[INFO] Cleaning report text...")
        hypotheses = [clean_report(h) for h in hypotheses_raw]
        references = [clean_report(r) for r in references_raw]
    else:
        hypotheses = hypotheses_raw
        references = references_raw

    n_total = len(references)
    n_valid = sum(1 for r, h in zip(references, hypotheses) if r.strip() and h.strip())
    print(f"[INFO] Evaluating {n_total} samples ({n_valid} non-empty pairs)")

    # 2. Compute metrics
    corpus_results = {}
    all_per_sample = [{} for _ in range(n_total)]
    timings = {}

    if "bleu" in metrics_to_compute:
        print(f"\n{'='*60}\n  Computing BLEU-1/2/3/4\n{'='*60}")
        t0 = time.time()
        bleu_results, bleu_per_sample = compute_bleu(references, hypotheses)
        timings["BLEU"] = time.time() - t0
        corpus_results.update(bleu_results)
        for i, ps in enumerate(bleu_per_sample):
            all_per_sample[i].update(ps)
        print(f"[RESULT] {bleu_results}")

    if "rouge" in metrics_to_compute:
        print(f"\n{'='*60}\n  Computing ROUGE-L\n{'='*60}")
        t0 = time.time()
        rouge_results, rouge_per_sample = compute_rouge(references, hypotheses)
        timings["ROUGE"] = time.time() - t0
        corpus_results.update(rouge_results)
        for i, ps in enumerate(rouge_per_sample):
            all_per_sample[i].update(ps)
        print(f"[RESULT] {rouge_results}")

    if "meteor" in metrics_to_compute:
        print(f"\n{'='*60}\n  Computing METEOR\n{'='*60}")
        t0 = time.time()
        meteor_results, meteor_per_sample = compute_meteor(references, hypotheses)
        timings["METEOR"] = time.time() - t0
        corpus_results.update(meteor_results)
        for i, ps in enumerate(meteor_per_sample):
            all_per_sample[i].update(ps)
        print(f"[RESULT] {meteor_results}")

    if "bertscore" in metrics_to_compute:
        print(f"\n{'='*60}\n  Computing BERTScore\n{'='*60}")
        t0 = time.time()
        bs_results, bs_per_sample = compute_bertscore(
            references, hypotheses,
            batch_size=args.bertscore_batch_size,
            model_type=args.bertscore_model,
        )
        timings["BERTScore"] = time.time() - t0
        corpus_results.update(bs_results)
        for i, ps in enumerate(bs_per_sample):
            all_per_sample[i].update(ps)
        print(f"[RESULT] {bs_results}")

    if "radgraph" in metrics_to_compute:
        print(f"\n{'='*60}\n  Computing RadGraph F1\n{'='*60}")
        t0 = time.time()
        rg_results, rg_per_sample = compute_radgraph_f1(
            references, hypotheses, model_type=args.radgraph_model
        )
        timings["RadGraph"] = time.time() - t0
        if rg_results is not None:
            corpus_results.update(rg_results)
            if rg_per_sample is not None:
                for i, ps in enumerate(rg_per_sample):
                    all_per_sample[i].update(ps)
            print(f"[RESULT] {rg_results}")

    if "ratescore" in metrics_to_compute:
        print(f"\n{'='*60}\n  Computing RaTEScore\n{'='*60}")
        t0 = time.time()
        rate_results, rate_per_sample = compute_ratescore(references, hypotheses)
        timings["RaTEScore"] = time.time() - t0
        if rate_results is not None:
            corpus_results.update(rate_results)
            if rate_per_sample is not None:
                for i, ps in enumerate(rate_per_sample):
                    all_per_sample[i].update(ps)
            print(f"[RESULT] {rate_results}")

    # 3. Print summary
    print(f"\n{'='*60}")
    print(f"  EVALUATION SUMMARY")
    print(f"{'='*60}")
    print(f"  Input:   {args.llm_output}")
    print(f"  Samples: {n_total} ({n_valid} non-empty)")
    print(f"{'─'*60}")
    for metric, value in corpus_results.items():
        print(f"  {metric:<25} {value:.4f}")
    print(f"{'─'*60}")
    print(f"  Timings:")
    for metric, t in timings.items():
        print(f"    {metric:<20} {t:.1f}s")
    print(f"{'='*60}")

    # 4. Save results
    xlsx_path = os.path.join(args.output_dir, f"{file_prefix}_nlg_metrics.xlsx")
    save_results_to_xlsx(corpus_results, args, n_total, xlsx_path)

    records = []
    for i, did in enumerate(dicom_ids):
        record = {"dicom_id": did}
        record.update(all_per_sample[i])
        records.append(record)

    per_sample_df = pd.DataFrame(records)
    csv_path = os.path.join(args.output_dir, f"{file_prefix}_nlg_per_sample.csv")
    per_sample_df.to_csv(csv_path, index=False)
    print(f"[INFO] Per-sample CSV saved to {csv_path}")

    json_path = os.path.join(args.output_dir, f"{file_prefix}_nlg_metrics.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "corpus_metrics": corpus_results,
            "n_samples": n_total,
            "n_valid": n_valid,
            "llm_output": args.llm_output,
            "metrics_computed": metrics_to_compute,
            "timings_seconds": timings,
        }, f, indent=2, ensure_ascii=False)
    print(f"[INFO] Metrics JSON saved to {json_path}")

    print(f"\n[DONE] NLG evaluation complete!")


if __name__ == "__main__":
    main()
